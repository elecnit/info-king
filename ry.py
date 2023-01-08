#import part of selenium
from re import search
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import pandas as pd
#import of openpyxl
import openpyxl as op
#import part of telegram Bot
from telegram import Bot 
from telegram.ext import Updater,CallbackContext,CommandHandler,MessageHandler
from telegram.update import Update
from telegram.ext.filters import Filters
import os


#About telegram Bot
bot = Bot('5914709645:AAHmR9YqqtBzgGwDjY29S8DPH8w83_ybcQY')
updater = Updater('5914709645:AAHmR9YqqtBzgGwDjY29S8DPH8w83_ybcQY')
dispater = updater.dispatcher

did_file_name = False
search_keyword = False
def start(update=Update,context=CallbackContext):
    bot.send_message(chat_id=update.effective_chat.id,text='hi,your talking to info King , please enter the any related things to job like place , job position , salary , technology etc  use this command to seach /search_create')
    search_create()
dispater.add_handler(CommandHandler('start',start))

def search_create(update=Update,context=CallbackContext):
    
    bot.send_message(chat_id=update.effective_chat.id,text='Enter the keyword to search job')
    def take_search_name(update=Update,context=CallbackContext):
        global OPO 
        global search_keyword
        global did_file_name
        if not search_keyword:
            keyword = update.message.text
            bot.send_message(chat_id=update.effective_chat.id,text='please wait your data is fetching!!!!!!')
            OPO = scrapy(keyword)
            if (OPO):
                search_keyword=True
                sleep(3)
            else:
                bot.send_message(chat_id=update.effective_chat.id,text='please enter the valid keyword')
        if not did_file_name:
            bot.send_message(chat_id=update.effective_chat.id,text='please enter the file name')
            if update.message.text:
                did_file_name = True
        else:
            xl_creater(OPO)
            file_name = update.message.text
            did_file_name = True
            xl_note.save(file_name+'.xlsx')
            bot.sendDocument(chat_id=update.effective_chat.id,document=open(file_name+'.xlsx','rb'))
            os.remove(file_name+'.xlsx') 
    dispater.add_handler(MessageHandler(Filters.text and (~Filters.command),take_search_name)) 
    
dispater.add_handler(CommandHandler('search_create',search_create))

# def file_name(update=Update,context=CallbackContext):
    
#     if not did_file_name:
#         bot.send_message(chat_id=update.effective_chat.id,text='please enter the file name')
#     else:
#         xl_creater(op)
#         file_name = update.message.text
#         did_file_name = True
#         xl_note.save(file_name+'.xlsx')
#         bot.sendDocument(chat_id=update.effective_chat.id,document=open(file_name+'.xlsx','rb'))
#         os.remove(file_name+'.xlsx')
# dispater.add_handler(CommandHandler('file_name',file_name))

#For Scraping 
def scrapy(value):
        path = "c://chromedriver.exe"
        browser = webdriver.Chrome(executable_path=path)
        browser.get('https://www.foundit.in/')
        input_search = browser.find_element('id','SE_home_autocomplete')
        input_search.send_keys(value)

        search_btn = browser.find_element('xpath',"//input[@class='btn']")
        search_btn.click()

        listy={}
        for i in range(5):
                        companies_info = browser.find_elements(By.XPATH,"//div[@class='cardContainer']")
                        for company in companies_info:
                            company_name = company.find_element(By.CLASS_NAME,"companyName")
                            com_name = company_name.find_element(By.TAG_NAME,'p').text
                            job_title = company.find_element(By.CLASS_NAME,"jobTitle").text
                            card = company.find_element(By.CLASS_NAME,"cardBody")
                            details = card.find_elements(By.CLASS_NAME,"details")

                            listy[com_name] = [job_title,[i.text for i in details]]
                        if  browser.find_element(By.CLASS_NAME,"mqfisrp-right-arrow"):
                          next = browser.find_element(By.CLASS_NAME,"mqfisrp-right-arrow")
                          next.click()
                        sleep(5)
    # print(com_name,job_title,if details[0].text, if details[1].text, if details[2].text)
   
        return listy

#data into xlsx formate
def xl_creater(opp):
    global xl_note
    output = opp
    xl_note = op.Workbook()
    xl_note_sheet = xl_note.active
    xl_note_sheet.cell(1,1).value = 'SL NO'
    xl_note_sheet.cell(1,2).value = 'Company Name'
    xl_note_sheet.cell(1,3).value = 'Job Title'
    xl_note_sheet.cell(1,4).value = 'full/half'
    xl_note_sheet.cell(1,5).value = 'Palce'
    xl_note_sheet.cell(1,6).value = 'Experience'
    xl_note_sheet.cell(1,7).value = 'salary'
    row = 2 
    col = 1
    for index,i in enumerate(output):
        xl_note_sheet.cell(row,col).value = index+1
        xl_note_sheet.cell(row,col+1).value = i
        xl_note_sheet.cell(row,col+2).value = output[i][0]
        for j in output[i][1]:
            xl_note_sheet.cell(row,col+3).value = j
            col+=1
        col=1
        row+=1
 
    

updater.start_polling()




        
        
